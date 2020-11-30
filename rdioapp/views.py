import http.client
from io import BytesIO
import json
import hmac
import codecs
import hashlib
import random
import urllib.parse
from datetime import datetime, date

import pandas as pd
import pyodbc
from django.conf import settings
from revproxy.views import ProxyView
from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule
from django.http import StreamingHttpResponse, HttpResponse
from rest_framework.decorators import api_view, authentication_classes
from rest_framework.response import Response
from rest_framework.authentication import SessionAuthentication
from rest_framework.serializers import ValidationError
from rest_framework.views import APIView

from .auth import Special_TokenAuthentication
from .models import Token

# VIEWS
@api_view(['GET','POST','PUT','PATCH','DELETE','OPTIONS'])
@authentication_classes([SessionAuthentication, Special_TokenAuthentication])
def proxy_requests(request):

    # if request.path.startswith("/rdio/_local/"):
    #     print('sleeping')
    #     time.sleep(10)

    the_token = request.META['HTTP_AUTHORIZATION'].split()[1]
    token_object = Token.objects.filter(key=the_token)[0]
    username = token_object.couch_user
    roles = str(token_object.couch_roles).replace('.',',')

    # Hash the secret and the username together
    secret = getattr(settings, 'COUCH_SECRET')
    digester = hmac.new(bytes(secret, 'UTF-8'), bytes(username, 'UTF-8'), hashlib.sha1)
    signature1 = digester.digest()
    hexlify = codecs.getencoder('hex')
    signature2 = hexlify(signature1)[0].decode('UTF-8')

    print('role123s', roles)
    print('username', username)

    # We start by building the headers
    headers = {
        "X-Auth-CouchDB-Token": signature2,
        'X-Auth-CouchDB-UserName': username,
        'X-Auth-CouchDB-Roles': roles,
        'Content-Type': 'application/json'
    }
    # We establish de connection
    conn = http.client.HTTPConnection(getattr(settings, 'COUCHDB_SERVER'))
    conn.request(request.method, request.get_full_path(), request.body, headers)
    the_response = conn.getresponse().read().decode('utf-8')

    return Response(json.loads(the_response))

class CustomProxyView(ProxyView):
    upstream = 'http://localhost:5984'

    def get_request_headers(self):
        # Call super to get default headers
        headers = super(CustomProxyView, self).get_request_headers()
        
        # getting token
        print(headers['Authorization'])
        sp=[]
        the_token = 'null'
        if headers['Authorization']:
            sp = headers['Authorization'].split()
        if len(sp) == 2:          
            the_token = sp[1]
        if not the_token == 'null':
            token_object = Token.objects.filter(key=the_token)[0]
            username = token_object.couch_user
            roles = str(token_object.couch_roles).replace('.',',')

            # Hash the secret and the username together
            secret = getattr(settings, 'COUCH_SECRET')
            digester = hmac.new(bytes(secret, 'UTF-8'), bytes(username, 'UTF-8'), hashlib.sha1)
            signature1 = digester.digest()
            hexlify = codecs.getencoder('hex')
            signature2 = hexlify(signature1)[0].decode('UTF-8')
            # Add new header

            headers['X-Auth-CouchDB-Token'] = signature2
            headers['X-Auth-CouchDB-UserName'] = username
            headers['X-Auth-CouchDB-Roles'] = roles
        #'Content-Type': 'application/json'
        return headers


def PerformAdminCouchOperation(couch_connection, headers, operation_method, request_path, body=''):
    try:
        couch_connection.request(operation_method, request_path, body ,headers)
        response = json.loads(couch_connection.getresponse().read().decode('utf-8'))
        # print(response)
    except Exception as e:
        print('Couch2 OPERATION error', e)
        response = 'nope'
    return response

def GetHapiQAViews(table, special_condition=''):
    # SQL Database connection
    conn = pyodbc.connect('DRIVER={SQL Server}; \
                        SERVER=' + getattr(settings, 'HAPIQA_SERVER') + '; \
                        DATABASE=' + getattr(settings, 'HAPIQA_DATABASE_NAME') + '; \
                        UID=' + getattr(settings, 'HAPIQA_DATABASE_USERNAME') + '; \
                        PWD=' + getattr(settings, 'HAPIQA_DATABASE_PASSWORD') + ';')
    cursor = conn.cursor()

    # Get SQL elements
    sql_view_name = getattr(settings, 'HAPIQA_VIEWS')[table]
    the_query = 'SELECT * FROM INMAC.dbo.' + sql_view_name + ' ' + special_condition + ';'
    cursor.execute(the_query)
    columns = [column[0] for column in cursor.description]
    sql_response = []
    for row in cursor.fetchall():
        sql_response.append(dict(zip(columns, row)))

    return sql_response  

def CreateUpdateObject(base_object, values_object):
    sql_keys = list(values_object.keys())
    update_object = base_object

    for key in sql_keys:
        update_object[key] = values_object[key]    

    return update_object 

def GetSpecialCompareObjects(sql_response, table):
    if table == 'proyecto':
        partidas = GetHapiQAViews('partidas')
        proyectos_main = {}
        for item in sql_response:
            proyectos_main[item['cod'].replace(' ','')] = item
            proyectos_main[item['cod'].replace(' ','')]['partidas'] = []
            proyectos_main[item['cod'].replace(' ','')]['cod'] = item['cod'].replace(' ','')
            if item['estado'] == True:
                item['estado'] = 'A'
            else:
                item['estado'] = 'I'

        for item in partidas:
            cod_proyecto = item['cod_proyecto'].replace(' ','')
            del item['cod_proyecto']
            item['cod'] = item['cod'].replace(' ','')
            proyectos_main[cod_proyecto]['partidas'].append(item)

    if table == 'persona':
        for item in sql_response:
            item['cod'] = item['cod'].replace(' ','')
            item['tipo'] = {
                'cod': item['tipo_cod'],
                'nombre': item['tipo_nombre']
            }
            del item['tipo_cod']
            del item['tipo_nombre']
            if item['estado'] == True:
                item['estado'] = 'A'
            else:
                item['estado'] = 'I'

    if table == 'pxe':
        for item in sql_response:
            item['proyecto'] = item['proyecto_cod'].replace(' ','')
            item['equipo'] = item['equipo_cod'].replace(' ','')
            item['categoria'] = {
                'cod': item['categoria_cod'],
                'nombre': item['categoria_nombre']
            }
            del item['categoria_cod']
            del item['categoria_nombre']        
            del item['proyecto_cod']        
            del item['equipo_cod']        

    if table == 'pxm':
        for item in sql_response:
            item['proyecto'] = item['proyecto_cod'].replace(' ','')
            item['categoria'] = {
                'cod': item['categoria_cod'],
                'nombre': item['categoria_nombre']
            }
            del item['categoria_cod']
            del item['categoria_nombre']        
            del item['proyecto_cod']        

    if table == 'pxp':
        for item in sql_response:
            item['proyecto'] = item['proyecto_cod'].replace(' ','')
            item['persona'] = item['persona_cod'].replace(' ','')
            item['categoria'] = {
                'cod': item['categoria_cod'],
                'nombre': item['categoria_nombre']
            }
            del item['categoria_cod']
            del item['categoria_nombre']        
            del item['proyecto_cod']     
            del item['persona_cod']     

    return sql_response

def CreateSpecialCouchID(item, table_element):
    variable = ''
    if table_element == 'pxe':
        variable = 'equipo:' + item['equipo'] + ':'
    if table_element == 'pxp':
        variable = 'persona:' + item['persona'] + ':'
    if table_element == 'pxm':
        variable = ''
    special_id = table_element + ':' + 'proyecto:' + item['proyecto'] + ':' + variable + 'categoria:' + item['categoria']['cod']
    return special_id

def CouchSqlCompareServer(table_list):
    operation_report = {}

    # Element validation
    valid_elements = getattr(settings, 'HAPIQA_VIEWS')
    # del valid_elements['partidas']
    for item in table_list:
        if item not in valid_elements:
            return {
                'status': 'error',
                'detail': str(item) + ' is not a valid element',
                'valid_elements': valid_elements.keys()
            }

    # Establish the main COUCH connection
    username = getattr(settings, 'COUCH_ADMIN_USER') 
    secret = getattr(settings, 'COUCH_SECRET')
    digester = hmac.new(bytes(secret, 'UTF-8'), bytes(username, 'UTF-8'), hashlib.sha1)
    signature1 = digester.digest()
    hexlify = codecs.getencoder('hex')
    signature2 = hexlify(signature1)[0].decode('UTF-8')

    headers = {
        "X-Auth-CouchDB-Token": signature2,
        'X-Auth-CouchDB-UserName': username,
        'X-Auth-CouchDB-Roles': '',
        'Content-Type': 'application/json'
    }
    couch_connection = http.client.HTTPConnection(CustomProxyView.upstream.replace('http://', ''))

    for table_element in table_list:
        print('Processing ' + table_element + '...')
        update_report = []
        creation_report = []
        simple_table_object = False

        # Get SQL views
        sql_response = GetHapiQAViews(table_element)

        # Get COUCH elements
        couch_request_path = '/' + getattr(settings, 'COUCH_DATABASE') + '/_all_docs?include_docs=true&inclusive_end=true&start_key=%22' + \
                table_element + '%3A%22&end_key=%22' + table_element + '%3Aufff0%22'
        couch_elements = PerformAdminCouchOperation(couch_connection, headers, 'GET', couch_request_path)

        # Index the Couch object dictionary
        couch_main_dict = {}
        for item in couch_elements['rows']:
            if table_element == 'pxe' or table_element == 'pxp' or table_element == 'pxm':
                special_identifier = CreateSpecialCouchID(item['doc'], table_element)
                couch_main_dict[special_identifier] = item['doc']
            if 'cod' in item['doc']:
                couch_main_dict[item['doc']['cod']] = item['doc']

        # SPECIAL COMPARE
        if table_element == 'persona' or table_element == 'proyecto' or table_element == 'pxe' or table_element == 'pxm' or table_element == 'pxp':
            sql_response = GetSpecialCompareObjects(sql_response, table_element)
        else:
            simple_table_object = True

        # Comparison
        sql_keys = list(sql_response[0].keys())
        for item in sql_response:
            if simple_table_object:
                if item['estado'] == True:
                    item['estado'] = 'A'
                if item['estado'] == False:
                    item['estado'] = 'I'
                item['cod'] = item['cod'].replace(' ','')                
            try:
                if table_element == 'pxe' or table_element == 'pxp' or table_element == 'pxm':
                    special_identifier = CreateSpecialCouchID(item, table_element)
                    couch_to_compare = couch_main_dict[special_identifier]
                else:
                    couch_to_compare = couch_main_dict[item['cod'].replace(' ','')]

                for key in sql_keys:
                    if item[key] != couch_to_compare[key] and key != 'cod':
                        # UPDATE OPERATION
                        update_object = CreateUpdateObject(couch_to_compare, item)
                        encoded_id = urllib.parse.quote(update_object['_id'])
                        update_path = '/' + getattr(settings, 'COUCH_DATABASE') + '/' + encoded_id
                        PerformAdminCouchOperation(couch_connection, headers, 'PUT', update_path, json.dumps(update_object).encode('utf-8'))
                        update_report.append('id ' + update_object['_id'] + ' updated successfully')
                        break

            except Exception as e:
                # CREATE OPERATION
                # print('e', e)
                create_object = item
                if table_element == 'pxe' or table_element == 'pxp' or table_element == 'pxm':
                    create_object['_id'] = CreateSpecialCouchID(create_object, table_element)
                if 'cod' in create_object:
                    create_object['_id'] = table_element + ':' + create_object['cod']
                if table_element == 'persona':
                    verification_field = random.randint(10000000, 99999999)
                    create_object['verificacion'] = verification_field
                create_path = '/' + getattr(settings, 'COUCH_DATABASE')
                PerformAdminCouchOperation(couch_connection, headers, 'POST', create_path, json.dumps(create_object).encode('utf-8'))
                creation_report.append('new id ' + create_object['_id'] + ' created successfully')

        # DELETE OPRATION
        # Index the SQL object dictionary
        delete_report = []
        sql_main_dict = {}
        for item in sql_response:
            if 'cod' in sql_response[0]:
                # sql_main_dict[item['cod']] = item
                sql_main_dict[table_element + ':' + item['cod']] = item
            if table_element == 'pxe' or table_element == 'pxp' or table_element == 'pxm':
                identifier = CreateSpecialCouchID(item, table_element)
                sql_main_dict[identifier] = item
                    
        for item in couch_elements['rows']:
            if not item['doc']['_id'] in sql_main_dict:
                encoded_id = urllib.parse.quote(item['doc']['_id'])
                delete_path = '/' + getattr(settings, 'COUCH_DATABASE') + '/' + encoded_id + '?rev=' + item['doc']['_rev']
                PerformAdminCouchOperation(couch_connection, headers, 'DELETE', delete_path)
                delete_report.append('id ' + item['doc']['_id'] + ' deleted successfully')

        operation_report[table_element] = {
            'UPDATE': update_report,
            'CREATE': creation_report,
            'DELETE': delete_report
            }

    return {
        "status": 200,
        "operation_report": operation_report
    }

@api_view(['POST'])
def CouchSqlCompare(request):

    # Parameter validation
    try:
        table_list = request.data['elements']
    except Exception as e:
        return Response({
            'status': 'error',
            'detail': "body should include the param 'elements'",
            'example': {
	                "elements": ["clima", "cuadrilla", "equipo"]
                    }
        })
    return Response(CouchSqlCompareServer(table_list))

def get_rdio(rdio_id):
    safe_rdio_id = urllib.parse.quote(rdio_id)

    # Establish the main COUCH connection
    username = getattr(settings, 'COUCH_ADMIN_USER') 
    secret = getattr(settings, 'COUCH_SECRET')
    digester = hmac.new(bytes(secret, 'UTF-8'), bytes(username, 'UTF-8'), hashlib.sha1)
    signature1 = digester.digest()
    hexlify = codecs.getencoder('hex')
    signature2 = hexlify(signature1)[0].decode('UTF-8')

    headers = {
        "X-Auth-CouchDB-Token": signature2,
        'X-Auth-CouchDB-UserName': username,
        'X-Auth-CouchDB-Roles': '',
        'Content-Type': 'application/json'
    }

    couch_connection = http.client.HTTPConnection(CustomProxyView.upstream.replace('http://', ''))

    # Get rdio data
    couch_request_path = '/' + getattr(settings, 'COUCH_DATABASE') + '/' + safe_rdio_id
    rdio = PerformAdminCouchOperation(couch_connection, headers, 'GET', couch_request_path)

    return rdio

@api_view(['POST'])
def TxtExport(request):
    rdio_id = request.data['rdio_id']
    safe_rdio_id = urllib.parse.quote(rdio_id)

    # Establish the main COUCH connection
    username = getattr(settings, 'COUCH_ADMIN_USER') 
    secret = getattr(settings, 'COUCH_SECRET')
    digester = hmac.new(bytes(secret, 'UTF-8'), bytes(username, 'UTF-8'), hashlib.sha1)
    signature1 = digester.digest()
    hexlify = codecs.getencoder('hex')
    signature2 = hexlify(signature1)[0].decode('UTF-8')

    headers = {
        "X-Auth-CouchDB-Token": signature2,
        'X-Auth-CouchDB-UserName': username,
        'X-Auth-CouchDB-Roles': '',
        'Content-Type': 'application/json'
    }
    
    couch_connection = http.client.HTTPConnection(CustomProxyView.upstream.replace('http://', ''))

    # Get rdio data
    couch_request_path = '/' + getattr(settings, 'COUCH_DATABASE') + '/' + safe_rdio_id
    rdio = PerformAdminCouchOperation(couch_connection, headers, 'GET', couch_request_path)
    old_partidas = rdio['partidas']
    partidas = {}

    for item in old_partidas:
        partidas[item['partida']['cod']] = item['partida']
        partidas[item['partida']['cod']]['n'] = item['n']
        partidas[item['partida']['cod']]['str_actividad'] = ''
        partidas[item['partida']['cod']]['avance'] = 0 if 'avance' not in item['partida'] else int(item['partida']['avance'])
        partidas[item['partida']['cod']]['actividades'] = []

        if 'actividades' in item: 
            for elem in item["actividades"]:
                partidas[item['partida']['cod']]['str_actividad'] += elem['title'] + ',' 

                partidas[item['partida']['cod']]['actividades'].append({
                    "n": item['n'],
                    "actividad": elem['title'],
                    "unidad": item['partida']['unidad']
                })
            partidas[item['partida']['cod']]['str_actividad'] = partidas[item['partida']['cod']]['str_actividad'][:-1]

    for item in partidas:
        if 'personal' not in partidas[item]:
            partidas[item]['personal'] = {}
        
        if 'personal' in rdio:
            for elem in rdio['personal']:
                for obj in elem['partidas']:
                    if obj['n'] == partidas[item]['n'] and obj['horas'] > 0:
                        if elem['cod'] not in partidas[item]['personal']:
                            partidas[item]['personal'][elem['cod']] = elem

        if 'equipos' not in partidas[item]:
            partidas[item]['equipos'] = {}
        
        if 'equipos' in rdio:
            for elem in rdio['equipos']:
                for obj in elem['partidas']:
                    if obj['n'] == partidas[item]['n'] and obj['horas'] > 0:
                        if elem['equipo']['cod'] not in partidas[item]['equipos']:
                            partidas[item]['equipos'][elem['equipo']['cod']] = elem

        if 'consumibles' not in partidas[item]:
            partidas[item]['consumibles'] = {}
        
        if 'consumibles' in rdio and len(rdio['consumibles']) > 0:
            for elem in rdio['consumibles']:
                for obj in elem['partidas']:
                    if elem['n'] == partidas[item]['n'] and int(obj['cantidad']) > 0:
                        if elem['consumible']['cod'] not in partidas[item]['consumibles']:
                            partidas[item]['consumibles'][elem['consumible']['cod']] = elem

    final_file = {}
    for item in partidas:
        counter = 1
        txt = []
        f_inicio = datetime.strptime(rdio['fecha'], "%Y-%m-%d").strftime("%d/%m/%Y %H:%M:%S")
        f_fin = datetime.strptime(rdio['fecha'], "%Y-%m-%d").strftime("%d/%m/%Y %H:%M:%S")
        # define default fields
        if 'refrigerio' not in rdio:
            rdio['refrigerio'] = 0
        if 'contratista' not in rdio:
            rdio['contratista'] = '20513250445'
        if 'supervision' not in rdio:
            rdio['supervision'] = '20513250445'
        clima = '01'
        if 'controlActividades' in rdio:
            for act in rdio['controlActividades']:
                if rdio['controlActividades'][act] == 'LL':
                    clima = '02'
        if 'inoperativo' not in rdio:
            # default random value
            rdio['inoperativo'] = '03'
        if 'valor_inoperativo' not in rdio:
            rdio['valor_inoperativo'] = ''
        if 'tipo' not in rdio:
            # default random value
            rdio['tipo'] = '001'
        if 'turno' not in rdio:
            # default random value
            rdio['turno'] = '01'
        if 'unidad' not in rdio:
            # default random value
            rdio['unidad'] = '001'
        if 'comentarios' not in rdio:
            rdio['comentarios'] = 'SIN COMENTARIOS'
        if 'unidades_procesadas' not in rdio:
            rdio['unidades_procesadas'] = 0
        

        cabecera_list = [
            'CABRDA',
            datetime.strptime(rdio['fecha'], "%Y-%m-%d").strftime("%d/%m/%Y"),
            # date.today().strftime("%d/%m/%Y"), 
            f_inicio,
            f_fin,
            str(rdio['refrigerio']), 
            rdio['contratista'],
            rdio['supervision'],
            rdio['rubro']['cod'],
            rdio['locacion']['cod'],
            rdio['sector']['cod'],
            clima,
            rdio['operativo']['cod'],
            rdio['cuadrilla']['cod'],
            rdio['inoperativo'],
            rdio['valor_inoperativo'],
            rdio['tipo'],
            rdio['turno'],
            rdio['unidad'],
            rdio['proyecto']['cod'],
            item,
            rdio['comentarios'],
            'LATITUD|LONGITUD',
            str(rdio['unidades_procesadas']),
            'Detalle de ' + partidas[item]['str_actividad'],
            'Actividad Próximo día',
            str(partidas[item]['avance']) + '//n',
        ]
        cabecera_string = '|'.join(cabecera_list)

        txt.append(cabecera_string)
        for elem in partidas[item]['personal']:
            if partidas[item]['personal'][elem]['categorias'][0]['cod'] == "null":
                partidas[item]['personal'][elem]['categorias'][0]['cod'] = ''
            if 'observacion' not in partidas[item]['personal'][elem]:
                partidas[item]['personal'][elem]['observacion'] = 'OBS01' 

            temp_list = [
                'PERRDA',
                str(counter),
                elem,
                partidas[item]['personal'][elem]['categorias'][0]['cod'],
                str(partidas[item]['personal'][elem]['horas']),
                partidas[item]['personal'][elem]['observacion'] + '//n'
            ]
            temp_string = '|'.join(temp_list)
            txt.append(temp_string)
            counter = counter + 1

        counter = 1
        for elem in partidas[item]['equipos']:
            if partidas[item]['equipos'][elem]['equipo']['categoria']['cod'] == "null":
                partidas[item]['equipos'][elem]['categoria']['cod'] = ''
            if 'observacion' not in partidas[item]['equipos'][elem]:
                partidas[item]['equipos'][elem]['observacion'] = 'OBS01' 


            temp_list = [
                'EQ1RDA',
                str(counter),
                partidas[item]['equipos'][elem]['equipo']['categoria']['cod'],
                elem,
                str(partidas[item]['equipos'][elem]['horometroInicial']),
                str(partidas[item]['equipos'][elem]['horometroFinal']),
                partidas[item]['equipos'][elem]['observacion'] + '//n'
            ]
            temp_string = '|'.join(temp_list)
            txt.append(temp_string)
            counter = counter + 1

        counter = 1
        for elem in partidas[item]['consumibles']:
            if not 'categoria' in partidas[item]['consumibles'][elem]:
                partidas[item]['consumibles'][elem]['categoria'] = {
                    "cod": ""
                }
            elif partidas[item]['consumibles'][elem]['categoria']['cod'] == "null":
                partidas[item]['consumibles'][elem]['categoria']['cod'] = ''
            if 'observacion' not in partidas[item]['consumibles'][elem]:
                partidas[item]['consumibles'][elem]['observacion'] = 'OBS01' 

            temp_list = [
                'MATRDA',
                str(counter),
                partidas[item]['consumibles'][elem]['consumible']['cod'],
                str(partidas[item]['consumibles'][elem]['partidas'][0]['cantidad']),
                partidas[item]['consumibles'][elem]['observacion'] + '//n'
            ]
            temp_string = '|'.join(temp_list)
            txt.append(temp_string)
            counter = counter + 1

        final_file[item] = txt

    return Response({
        'status': 200,
        'detail': final_file
    })

class BasicReport(APIView):
    def post(self, request):
        try:
            rdio = get_rdio(request.data['rdio_id'])
        except Exception as e:
            return Response(e)

        # [PARTIDAS]
        partidas = [] if not 'partidas' in rdio else rdio['partidas']
        partidas_list = []
        for item in partidas:
            p = {
                "Cod": item['partida']['cod'],
                "Nombre": item['partida']['nombre'],
                "Descripcion": item['partida']['descripcion'],
                'Unidad': item['partida']['unidad']['nombre'],
                'Metrado': item['cantidadTotal'],
                'Porcentaje Avance': None if 'metrado_total' not in item['partida'] \
                                     else float(item['cantidadTotal']) / item['partida']['metrado_total'],
                'Metrado Anterior': 0,
                'Metrado Total': None if 'metrado_total' not in item['partida'] \
                                 else item['partida']['metrado_total']
            }
            partidas_list.append(p)

        partidas_df = pd.DataFrame(partidas_list)
        partidas_df['Metrado'] = partidas_df['Metrado'].astype(float)

        # [PERSONAL]
        personal = [] if not 'personal' in rdio else rdio['personal']
        personal_list = []
        for item in personal:
            item['Cargo'] = item['categoria']['nombre']
            item['Total Horas'] = 0
            for elem in item['partidas']:
                item['Total Horas'] += elem["horas"]
            personal_list.append(item)

        personal_df = pd.DataFrame(personal_list)
        personal_df.rename(columns={'nombre': 'Nombre', 'dni': 'DNI'}, errors='ignore', inplace=True)
        personal_df = personal_df.loc[:, personal_df.columns.intersection(['DNI','Nombre','Cargo','Total Horas'])]

        # [EQUIPOS]
        equipos = [] if not 'equipos' in rdio else rdio['equipos']
        equipos_list = []
        for item in equipos:
            equipos_list.append(
                {
                    'Cod': item['equipo']['cod'],
                    'Nombre': item['equipo']['nombre'],
                    'Cantidad': int(item['cantidad']),
                    'Horometro Inicial': None if 'horometroInicial' not in item else int(item['horometroInicial']),
                    'Horometro Final': None if 'horometroFinal' not in item else int(item['horometroFinal']),
                    'Horas Totales': None if 'horometroInicial' not in item and 'horometroFinal' not in item else int(item['horometroInicial']) + int(item['horometroFinal'])
                }
            )
        
        equipos_df = pd.DataFrame(equipos_list)
        equipos_df = equipos_df.loc[:, equipos_df.columns.intersection(['Cod','Nombre','Cantidad','Horometro Inicial', 'Horometro Final', 'Horas Totales'])]

        # [CONSUMIBLES]
        consumibles = [] if not 'consumibles' in rdio else rdio['consumibles']
        consumibles_list = []
        for item in consumibles:
            temp_elem = {}
            temp_elem = {
                "Cod": item['consumible']['cod'],
                'Nombre': item['consumible']['nombre'],
                'Unidad': item['unidad']['nombre'],
                'Observacion': item['observacion'],
                'Horas': 0
                }

            if 'partidas' in item:
                for elem in item['partidas']:
                    temp_elem['Horas'] += elem['horas']
            consumibles_list.append(temp_elem)

        consumibles_df = pd.DataFrame(consumibles_list)



        # get all previous RDIOs
        prev_rdios = self.get_prev_rdios(rdio['fecha'], rdio['proyecto']['cod'], True)['docs']
        
        prev_partida_progress = {}
        for r in prev_rdios:
            if 'partidas' in r:
                for partida in r['partidas']:
                    if partida['partida']['cod'] not in prev_partida_progress:
                        prev_partida_progress[partida['partida']['cod']] = partida['cantidadTotal']
                    else:
                        prev_partida_progress[partida['partida']['cod']] += partida['cantidadTotal']

        partidas_df['Metrado Anterior'] = partidas_df["Cod"].apply(lambda x: prev_partida_progress.get(x))
        partidas_df['Metrado Anterior'] = partidas_df['Metrado Anterior'].astype(float)
        partidas_df['Metrado Ac. Total'] = partidas_df['Metrado'] + partidas_df['Metrado Anterior'].astype(float)

        partidas_df['Avance Total'] = None
        partidas_df.loc[partidas_df['Metrado Total'] != None, 'Avance Total'] = partidas_df['Metrado Ac. Total'] \
                                                                                  / partidas_df['Metrado Total'] 
        partidas_df.drop(columns=['Metrado Total'], errors='ignore', axis=1, inplace=True)


        # [FOR TESTING]
        # self.test_report_existing_excel(
        #     [
        #         (partidas_df, 'PARTIDAS'),
        #         (personal_df, 'PERSONAL'),
        #         (equipos_df, 'EQUIPOS'),
        #         (consumibles_df, 'CONSUMIBLES')
        #     ], rdio)
        # return Response('Done!')

        # [PRODUCCION]
        file_name = 'reporte_rdio_%s.xlsx' % (rdio['fecha'])
        workbook = self.export_dataframes([(partidas_df, 'PARTIDAS'), 
                                          (personal_df, 'PERSONAL'), 
                                          (equipos_df, 'EQUIPOS'), 
                                          (consumibles_df, 'CONSUMIBLES')
                                          ], rdio)
        content_type1 = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # content_type2 = 'application/force-download'
        
        response = HttpResponse(workbook, content_type=content_type1)
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name

        return response

    def test_report_existing_excel(self, df_list, rdio):
        '''
        Para testing: Guarda el excel en la carpeta del proyecto
        '''
        xlsx_file = 'test3.xlsx'
        book = load_workbook(xlsx_file, data_only=True)
        writer = pd.ExcelWriter(xlsx_file, engine='openpyxl')
        writer.book = book

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        ws = writer.sheets['rdio']
        self.add_project_info(ws, rdio)
        self.write_tables(ws, writer, df_list)

        writer.save()


    def export_dataframes(self, df_list, rdio):
        '''
        Para producción. Toma el template del excel en el proyecto y contruye
        uno nuevo en memoria para enviarlo.
        '''
        sio = BytesIO()

        row = 8
        xlsx_file = settings.REPORT_TEMPLATE
        book = load_workbook(xlsx_file, data_only=True)
        writer = pd.ExcelWriter(sio, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        ws = writer.sheets['rdio']
        
        self.add_project_info(ws, rdio)
        self.write_tables(ws, writer, df_list)

        writer.save()
        sio.seek(0)
        return sio.getvalue()


    def write_tables(self, ws, writer, df_list, spaces=2):
        '''
        escribe las tablas en el template proporcionado
        '''
        row = 8
        for df in df_list:
            ws.insert_rows(idx=row+1 ,amount=len(df[0].index))
            df[0].to_excel(writer, sheet_name="rdio", startrow=row, startcol=0, header=False)

            if df[1] == 'PARTIDAS':
                rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                                color='FFDD00', showValue="None", minLength=None, maxLength=None)
                cells = 'J9:J%s' % (str(9+len(df[0].index)))
                ws.conditional_formatting.add(cells, rule)
                for i in range(0, len(df[0].index)):
                    ws['J' + str(9+i)].style = "Percent"
                    ws['G' + str(9+i)].style = "Percent"

            row = row + len(df[0].index) + spaces + 2


    def add_project_info(self, ws, rdio):
        ws['C2'] = rdio['proyecto']['cod']
        ws['C1'] = rdio['proyecto']['nombre']
        ws['C3'] = rdio['fecha']
        ws['C4'] = rdio['username']
        ws['H2'] = rdio['locacion']['nombre']
        ws['H3'] = rdio['rubro']['nombre']
        ws['H4'] = rdio['sector']['nombre']
        ws['N2'] = rdio['operativo']['nombre']
        ws['N3'] = rdio['cuadrilla']['nombre']


    def get_prev_rdios(self, date, p_code, only_partidas=False):

        # Establish the main COUCH connection
        username = getattr(settings, 'COUCH_ADMIN_USER') 
        secret = getattr(settings, 'COUCH_SECRET')
        digester = hmac.new(bytes(secret, 'UTF-8'), bytes(username, 'UTF-8'), hashlib.sha1)
        signature1 = digester.digest()
        hexlify = codecs.getencoder('hex')
        signature2 = hexlify(signature1)[0].decode('UTF-8')

        headers = {
            "X-Auth-CouchDB-Token": signature2,
            'X-Auth-CouchDB-UserName': username,
            'X-Auth-CouchDB-Roles': '',
            'Content-Type': 'application/json'
        }
        body = {
            "selector": {
                "_id": {
                    "$regex": p_code
                },
                "fecha": {
                    "$lt": date
                }
            }
        }
        if only_partidas:
            body["fields"] = ["partidas"]
        
        couch_connection = http.client.HTTPConnection(CustomProxyView.upstream.replace('http://', ''))

        # Get rdio data
        couch_request_path = '/' + getattr(settings, 'COUCH_DATABASE') + '/_find'
        prev_rdios = PerformAdminCouchOperation(couch_connection, headers, 'POST', couch_request_path, json.dumps(body).encode('utf-8'))

        return prev_rdios


